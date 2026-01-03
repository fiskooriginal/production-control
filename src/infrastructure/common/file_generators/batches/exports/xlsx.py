import asyncio

from dataclasses import fields
from functools import partial
from io import BytesIO

from openpyxl import Workbook

from src.application.batches.dtos.raw_data import BatchRawDataDTO
from src.application.batches.mappers import raw_data_dto_to_row
from src.core.logging import get_logger
from src.infrastructure.common.exceptions.batches import BatchExcelGenerationError

logger = get_logger("batches.export.xlsx")


class BatchesExportExcelGenerator:
    """Генератор Excel файлов для экспорта списка партий."""

    async def generate(self, raw_data: list[BatchRawDataDTO]) -> bytes:
        """
        Генерирует Excel файл со списком партий.

        Args:
            raw_data: Список партий для экспорта

        Returns:
            Байты сгенерированного Excel файла

        Raises:
            BatchExcelGenerationError: При ошибке генерации Excel
        """
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, partial(self._generate_sync, raw_data))

    def _generate_sync(self, batches: list[BatchRawDataDTO]) -> bytes:
        """
        Синхронная генерация Excel файла.

        Args:
            batches: Список партий для экспорта

        Returns:
            Байты сгенерированного Excel файла

        Raises:
            BatchExcelGenerationError: При ошибке генерации Excel
        """
        try:
            wb = Workbook()
            ws = wb.active

            headers = [field.name for field in fields(BatchRawDataDTO)]

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header

            row = 2
            for batch in batches:
                row_values = raw_data_dto_to_row(batch)
                for col_idx, value in enumerate(row_values, start=1):
                    ws.cell(row=row, column=col_idx).value = value
                row += 1

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            result_bytes = buffer.getvalue()

            logger.info(f"Сгенерирован Excel файл для экспорта {len(batches)} партий")
            return result_bytes
        except Exception as e:
            logger.error(f"Ошибка генерации Excel файла для экспорта: {e}")
            raise BatchExcelGenerationError(f"Ошибка генерации Excel файла для экспорта: {e}") from e
