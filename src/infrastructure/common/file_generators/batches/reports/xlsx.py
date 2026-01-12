import asyncio

from functools import partial
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.application.batches.reports.dtos import BatchReportDataDTO
from src.core.logging import get_logger
from src.infrastructure.common.exceptions.batches import BatchExcelGenerationError

logger = get_logger("batches.reports.excel")


class BatchExcelReportGenerator:
    """Генератор Excel отчетов для партий."""

    async def generate(self, batch_data: BatchReportDataDTO) -> bytes:
        """
        Генерирует Excel отчет для партий.

        Args:
            batch_data: Данные для генерации отчета

        Returns:
            Байты сгенерированного Excel файла

        Raises:
            BatchExcelGenerationError: При ошибке генерации Excel
        """
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(None, partial(self._generate_sync, batch_data))

    def _generate_sync(self, batch_data: BatchReportDataDTO) -> bytes:
        """
        Синхронная генерация Excel отчета.

        Args:
            batch_data: Данные для генерации отчета

        Returns:
            Байты сгенерированного Excel файла

        Raises:
            BatchExcelGenerationError: При ошибке генерации Excel
        """
        try:
            if not batch_data:
                raise BatchExcelGenerationError("batch_data не может быть None")

            wb = Workbook()
            wb.remove(wb.active)

            self._create_batch_info_sheet(wb, batch_data)
            self._create_products_sheet(wb, batch_data)
            self._create_statistics_sheet(wb, batch_data)

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            result_bytes = buffer.getvalue()

            logger.info(f"Сгенерирован Excel отчет для батча {batch_data.batch.batch_number}")
            return result_bytes
        except Exception as e:
            logger.error(f"Ошибка генерации Excel отчета: {e}")
            raise BatchExcelGenerationError(f"Ошибка генерации Excel отчета: {e}") from e

    @staticmethod
    def _create_batch_info_sheet(wb: Workbook, batch_data: BatchReportDataDTO) -> None:
        """Создает лист с информацией о партии."""
        ws = wb.create_sheet("Информация о партии", 0)

        batch = batch_data.batch

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        alignment_left = Alignment(horizontal="left", vertical="center")

        ws["A1"] = "Параметр"
        ws["B1"] = "Значение"

        header_cells = [ws["A1"], ws["B1"]]
        for cell in header_cells:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment_left

        row = 2
        data = [
            ("Номер партии", str(batch.batch_number)),
            ("Дата партии", batch.batch_date.strftime("%d.%m.%Y")),
            ("Номенклатура", str(batch.nomenclature)),
            ("Код ЕКН", str(batch.ekn_code)),
            ("Смена", str(batch.shift)),
            ("Бригада", str(batch.team)),
        ]

        if batch_data.work_center_name:
            data.append(("Рабочий центр", str(batch_data.work_center_name)))

        data.extend(
            [
                ("Описание задачи", str(batch.task_description)),
                ("Начало смены", batch.shift_time_range.start.strftime("%d.%m.%Y %H:%M")),
                ("Конец смены", batch.shift_time_range.end.strftime("%d.%m.%Y %H:%M")),
                ("Статус", "Закрыта" if batch.is_closed else "Открыта"),
            ]
        )

        if batch.closed_at:
            data.append(("Дата закрытия", batch.closed_at.strftime("%d.%m.%Y %H:%M")))

        for key, value in data:
            ws[f"A{row}"] = key
            ws[f"B{row}"] = value

            ws[f"A{row}"].font = header_font
            ws[f"A{row}"].border = border
            ws[f"A{row}"].alignment = alignment_left
            ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

            ws[f"B{row}"].border = border
            ws[f"B{row}"].alignment = alignment_left

            row += 1

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 50

    @staticmethod
    def _create_products_sheet(wb: Workbook, batch_data: BatchReportDataDTO) -> None:
        """Создает лист с продукцией."""
        ws = wb.create_sheet("Продукция", 1)

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=11, color="FFFFFF")
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        alignment_center = Alignment(horizontal="center", vertical="center")
        alignment_left = Alignment(horizontal="left", vertical="center")

        headers = ["ID", "Уникальный код", "Аггрегирована", "Дата аггрегации"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment_center

        row = 2
        for product in batch_data.batch.products:
            ws.cell(row=row, column=1).value = str(product.uuid)
            ws.cell(row=row, column=2).value = str(product.unique_code)

            aggregated_value = "Да" if product.is_aggregated else "Нет"
            ws.cell(row=row, column=3).value = aggregated_value
            ws.cell(row=row, column=4).value = (
                product.aggregated_at.strftime("%d.%m.%Y %H:%M") if product.aggregated_at else ""
            )

            for col_idx in range(1, 5):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = border
                cell.alignment = alignment_center if col_idx == 3 else alignment_left

            row += 1

        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 20

    @staticmethod
    def _create_statistics_sheet(wb: Workbook, batch_data: BatchReportDataDTO) -> None:
        """Создает лист со статистикой."""
        ws = wb.create_sheet("Статистика", 2)

        stats = batch_data.statistics

        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")
        border = Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )
        alignment_left = Alignment(horizontal="left", vertical="center")
        alignment_right = Alignment(horizontal="right", vertical="center")

        ws["A1"] = "Параметр"
        ws["B1"] = "Значение"

        header_cells = [ws["A1"], ws["B1"]]
        for cell in header_cells:
            cell.font = header_font_white
            cell.fill = header_fill
            cell.border = border
            cell.alignment = alignment_left

        data = [
            ("Всего продукции", stats.total_products),
            ("Аггрегировано", stats.aggregated_products),
            ("Осталось", stats.remaining_products),
            ("Процент выполнения", stats.completion_percentage),
            ("Средняя скорость (продуктов/час)", stats.average_speed),
        ]

        row = 2
        for key, value in data:
            ws[f"A{row}"] = key
            cell_b = ws[f"B{row}"]

            if key == "Процент выполнения":
                cell_b.value = stats.completion_percentage / 100
                cell_b.number_format = "0.00%"
            elif key == "Средняя скорость (продуктов/час)":
                cell_b.value = stats.average_speed
                cell_b.number_format = "0.00"
            else:
                cell_b.value = value
                cell_b.number_format = "General"

            ws[f"A{row}"].font = header_font
            ws[f"A{row}"].border = border
            ws[f"A{row}"].alignment = alignment_left
            ws[f"A{row}"].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

            cell_b.border = border
            cell_b.alignment = alignment_right

            row += 1

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 20
