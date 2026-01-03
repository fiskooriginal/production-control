import asyncio

from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from src.application.common.exceptions.file_parser import FileParseError
from src.application.common.ports.file_parser import FileParserProtocol


class XlsxFileParser(FileParserProtocol):
    """
    Парсер для XLSX файлов.

    Использует openpyxl для чтения Excel файлов.
    Работает в асинхронном режиме через executor для блокирующих операций.
    """

    async def parse(self, file_data: bytes) -> list[dict[str, Any]]:
        """
        Парсит XLSX файл и возвращает список словарей с данными.

        Args:
            file_data: Байты XLSX файла

        Returns:
            Список словарей, где каждый словарь представляет строку данных.
            Ключи словаря - это значения из первой строки (заголовки).

        Raises:
            FileParseError: При ошибке парсинга файла
        """
        try:
            return await asyncio.to_thread(self._parse_sync, file_data)
        except FileParseError:
            raise
        except Exception as e:
            raise FileParseError(f"Ошибка парсинга XLSX файла: {e}") from e

    def _parse_sync(self, file_data: bytes) -> list[dict[str, Any]]:
        """
        Синхронный метод парсинга XLSX файла.

        Args:
            file_data: Байты XLSX файла

        Returns:
            Список словарей с данными

        Raises:
            FileParseError: При ошибке парсинга
        """
        try:
            file_stream = BytesIO(file_data)
            workbook = load_workbook(file_stream, read_only=True, data_only=True)

            if not workbook.sheetnames:
                workbook.close()
                raise FileParseError("XLSX файл не содержит листов")

            worksheet = workbook.active

            rows = list(worksheet.iter_rows(values_only=True))

            if not rows:
                workbook.close()
                return []

            headers = [str(cell) if cell is not None else "" for cell in rows[0]]

            if not headers or all(not header for header in headers):
                workbook.close()
                raise FileParseError("XLSX файл не содержит заголовков в первой строке")

            data_rows = rows[1:]

            result = []
            for row in data_rows:
                if all(cell is None for cell in row):
                    continue

                row_dict = {}
                for idx, header in enumerate(headers):
                    value = row[idx] if idx < len(row) else None
                    row_dict[header] = value

                result.append(row_dict)

            workbook.close()
            return result

        except FileParseError:
            raise
        except Exception as e:
            raise FileParseError(f"Ошибка чтения XLSX файла: {e}") from e
